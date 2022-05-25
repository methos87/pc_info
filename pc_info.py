#!usr/bin/python

import platform
import socket
import re
import uuid
import psutil
from openpyxl import load_workbook


def get_System_Info():

        pf_system = platform.system()
        pf_release = platform.release()
        pf_version = platform.version()
        pf_machine = platform.machine()
        hostname = socket.gethostname()
        ip = socket.gethostbyname(socket.gethostname())
        mac_add = ':'.join(re.findall('..', '%012x' % uuid.getnode()))
        proc = platform.processor()
        ram = str(round(psutil.virtual_memory().total / (1024.0 **3)))+" GB"
        
        print(pf_system, pf_release, pf_version, pf_machine, hostname, ip, mac_add, proc, ram)
        
        return pf_system, pf_release, pf_version, pf_machine, hostname, ip, mac_add, proc, ram
        
def export_excel(pf_system, pf_release, pf_version, pf_machine, hostname, ip, mac_add, proc, ram):

        print("Exporting to excel....")
        
        workbook = load_workbook(filename="/storage/emulated/0/_python/excel/pc_info.xlsx")
        sheet = workbook.active
        
        for row in sheet.iter_cols(min_row=1,
                                                        max_row=9,
                                                        min_col=2,
                                                        max_col=2,
                                                        values_only=True):
              if str(mac_add) != str(row[6]):
                   sheet.insert_cols(idx=2)
                   
                   sheet["B1"] = str(pf_system)
                   sheet["B2"] = str(pf_release)
                   sheet["B3"] = str(pf_version)
                   sheet["B4"] = str(pf_machine)
                   sheet["B5"] = str(hostname)
                   sheet["B6"] = str(ip)
                   sheet["B7"] = str(mac_add)
                   sheet["B8"] = str(proc)
                   sheet["B9"] = str(ram)
                   
                   print("New elements inserted!")

              else:
                    print("This machine is already added!")                    
        
        workbook.save(filename = "/storage/emulated/0/_python/excel/pc_info.xlsx")

        print("Export is finished!")


export_excel(*get_System_Info())
